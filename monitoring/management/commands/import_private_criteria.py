from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from monitoring.models import InspectionType, CriterionCategory, Criterion


class Command(BaseCommand):
    help = "Импорт критериев для проверки частных школ из файла +по ЧШ.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Путь к файлу +по ЧШ.xlsx (например, '+по ЧШ.xlsx')",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]

        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Не удалось открыть файл {excel_path}: {e}")

        if "Лист1" not in wb.sheetnames:
            raise CommandError("В файле нет листа 'Лист1'")

        ws = wb["Лист1"]

        header_row = 10
        subheader_row = 11
        max_col = 39  # до 39-го столбца есть что-то осмысленное

        row10 = [ws.cell(row=header_row, column=j).value for j in range(1, max_col + 1)]
        row11 = [ws.cell(row=subheader_row, column=j).value for j in range(1, max_col + 1)]

        # 1) Вид проверки
        inspection_type, _ = InspectionType.objects.get_or_create(
            code="private_schools",
            defaults={
                "name": "Проверка частных школ",
                "description": "Критерии из шаблона +по ЧШ.xlsx",
                "is_active": True,
            },
        )

        # 2) Разделы критериев
        cat_gz_lc, _ = CriterionCategory.objects.get_or_create(
            inspection_type=inspection_type,
            name="Госзаказ и лицензия",
            defaults={"order": 1},
        )
        cat_km, _ = CriterionCategory.objects.get_or_create(
            inspection_type=inspection_type,
            name="Контингент и мощность",
            defaults={"order": 2},
        )
        cat_dir, _ = CriterionCategory.objects.get_or_create(
            inspection_type=inspection_type,
            name="Управление и кадровый состав",
            defaults={"order": 3},
        )
        cat_mto, _ = CriterionCategory.objects.get_or_create(
            inspection_type=inspection_type,
            name="Материально-техническое обеспечение",
            defaults={"order": 4},
        )

        def choose_category(col_index: int) -> CriterionCategory:
            """
            Колонки в файле логически разбиты на блоки,
            маппим их на 4 раздела.
            """
            if col_index in (2, 3):
                return cat_gz_lc
            elif 4 <= col_index <= 8:
                return cat_km
            elif 9 <= col_index <= 23:
                return cat_dir
            elif 24 <= col_index <= max_col:
                return cat_mto
            return cat_mto

        def infer_value_type(name: str) -> str:
            """
            По формулировке определяем тип значения.
            Используем те же коды, что в Criterion.ValueType.
            """
            text = name.lower()
            if "да-1/нет-0" in text or "да - 1/нет - 0" in text:
                return Criterion.ValueType.BOOLEAN
            if any(word in text for word in ["количество", "кол-во", "мощность", "контингент", "уч-ся", "учащ", "всего педагогов"]):
                return Criterion.ValueType.INTEGER
            if "стаж" in text:
                return Criterion.ValueType.INTEGER
            return Criterion.ValueType.TEXT

        created = 0
        updated = 0

        # будем накапливать порядковые номера по разделам
        order_by_category_id = {}

        current_group_title = None

        for col in range(2, max_col + 1):
            base_title = row10[col - 1]
            sub_title = row11[col - 1]

            # обновляем "группу" (верхний заголовок) если он есть
            if base_title:
                current_group_title = str(base_title).strip()

            if base_title is None and sub_title is None:
                continue

            # формируем полное название критерия
            if sub_title and current_group_title:
                full_name = f"{current_group_title.strip()} — {str(sub_title).strip()}"
            elif base_title:
                full_name = str(base_title).strip()
            else:
                full_name = str(sub_title).strip()

            category = choose_category(col)
            value_type = infer_value_type(full_name)

            # порядковый номер внутри категории
            cat_id = category.id or f"tmp-{category.name}"
            order_by_category_id.setdefault(cat_id, 0)
            order_by_category_id[cat_id] += 1
            order = order_by_category_id[cat_id]

            # простой код критерия по номеру колонки
            code = f"COL_{col}"

            criterion, created_flag = Criterion.objects.update_or_create(
                category=category,
                code=code,
                defaults={
                    "name": full_name,
                    "value_type": value_type,
                    "max_score": None,
                    "order": order,
                    "is_active": True,
                },
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"Импорт критериев завершён. Создано: {created}, обновлено: {updated}."
        ))
