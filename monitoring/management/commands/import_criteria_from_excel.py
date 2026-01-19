from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from monitoring.models import InspectionType, CriterionCategory, Criterion


def clean(v):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def infer_value_type(label: str) -> str:
    """
    Грубая, но рабочая эвристика под твой шаблон:
    - дата -> date
    - да/нет -> bool
    - количество/всего/кол-во/стаж -> int
    - иначе -> text
    """
    s = (label or "").lower()

    if "дата" in s:
        return "date"   # если добавил ValueType.DATE, иначе верни "text"

    if ("да-1/нет-0" in s) or ("да - 1/нет - 0" in s) or ("(да-1/нет-0" in s) or ("(да-1" in s and "нет" in s):
        return "bool"

    int_markers = ["количество", "кол-во", "всего", "стаж", "кабинет", "педагог", "уч-ся", "учащ"]
    if any(m in s for m in int_markers):
        return "int"

    return "text"


class Command(BaseCommand):
    help = "Импорт критериев (группы и подкритерии) из Excel-шаблона."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Путь к Excel файлу .xlsx")
        parser.add_argument("--sheet", default=None, help="Имя листа (если не указать — active)")
        parser.add_argument("--inspection-type-code", required=True, help="Код вида проверки (unique)")
        parser.add_argument("--inspection-type-name", required=True, help="Название вида проверки")
        parser.add_argument("--category-name", default="Таблица (Excel)", help="Название категории (одна на шаблон)")
        parser.add_argument("--header-row", type=int, default=10, help="Строка верхних заголовков (красные)")
        parser.add_argument("--subheader-row", type=int, default=11, help="Строка подзаголовков (синие)")
        parser.add_argument("--reset", action="store_true", help="Удалить старые критерии этой категории и создать заново")
        parser.add_argument("--dry-run", action="store_true", help="Только показать что будет создано, без записи в БД")

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts["file"]
        sheet_name = opts["sheet"]
        header_row = opts["header_row"]
        subheader_row = opts["subheader_row"]
        dry_run = opts["dry_run"]
        reset = opts["reset"]

        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name else wb.active

        max_col = ws.max_column

        # Список merged ranges для поиска “группы” по колонке
        merged = list(ws.merged_cells.ranges)

        def top_header_for_col(col: int) -> str:
            v = ws.cell(header_row, col).value
            if v is not None:
                return clean(v)

            # если ячейка пустая из-за merge — ищем range
            for r in merged:
                if r.min_row <= header_row <= r.max_row and r.min_col <= col <= r.max_col:
                    return clean(ws.cell(r.min_row, r.min_col).value)
            return ""

        def is_group_col(col: int) -> bool:
            """
            True если над колонкой есть merge по строке header_row шире 1 колонки.
            """
            for r in merged:
                if r.min_row == header_row and r.max_row == header_row and r.min_col <= col <= r.max_col:
                    return (r.max_col - r.min_col) >= 1
            return False

        # --- Собираем структуру ---
        # items: [{col, group_name, leaf_name, is_group}]
        items = []
        for col in range(1, max_col + 1):
            group_name = top_header_for_col(col)
            sub = clean(ws.cell(subheader_row, col).value)

            if not group_name and not sub:
                continue

            group = is_group_col(col)
            leaf_name = sub if sub else group_name

            items.append(
                {
                    "col": col,
                    "group_name": group_name,
                    "leaf_name": leaf_name,
                    "has_parent": bool(sub and group),
                }
            )

        # Покажем что нашли
        self.stdout.write(self.style.SUCCESS(f"Лист: {ws.title} | колонок: {max_col} | найдено полей: {len(items)}"))
        for it in items:
            if it["has_parent"]:
                self.stdout.write(f"  C{it['col']}: [{it['group_name']}] -> {it['leaf_name']}")
            else:
                self.stdout.write(f"  C{it['col']}: {it['leaf_name']}")

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: ничего не записано в БД."))
            return

        # --- Создаём/обновляем справочники ---
        itype, _ = InspectionType.objects.get_or_create(
            code=opts["inspection_type_code"],
            defaults={"name": opts["inspection_type_name"], "is_active": True},
        )
        # если имя поменялось — обновим
        if itype.name != opts["inspection_type_name"]:
            itype.name = opts["inspection_type_name"]
            itype.save(update_fields=["name"])

        category, _ = CriterionCategory.objects.get_or_create(
            inspection_type=itype,
            name=opts["category_name"],
            defaults={"order": 0},
        )

        if reset:
            Criterion.objects.filter(category=category).delete()
            self.stdout.write(self.style.WARNING("RESET: старые критерии категории удалены."))

        # --- Родители по группам ---
        parents = {}  # group_name -> Criterion
        created_parents = 0
        created_children = 0
        updated_children = 0

        # порядок родителей по первому появлению
        parent_order = 0

        for it in items:
            if not it["has_parent"]:
                continue

            gname = it["group_name"]
            if not gname:
                continue

            if gname not in parents:
                parent_order += 1
                obj, created = Criterion.objects.get_or_create(
                    category=category,
                    parent=None,
                    name=gname,
                    defaults={
                        "code": f"G{it['col']}",
                        "value_type": "text",
                        "order": parent_order,
                        "is_active": True,
                        "excel_col": None,
                    },
                )
                if created:
                    created_parents += 1
                parents[gname] = obj

        # --- Дети / одиночные колонки ---
        # order внутри одного parent по колонке
        child_order_map = {}

        for it in items:
            col = it["col"]
            name = it["leaf_name"]
            parent = parents.get(it["group_name"]) if it["has_parent"] else None

            key = parent.id if parent else 0
            child_order_map[key] = child_order_map.get(key, 0) + 1
            order = child_order_map[key]

            defaults = {
                "code": f"C{col}",
                "value_type": infer_value_type(name),
                "order": order,
                "is_active": True,
                "excel_col": col,
            }

            obj, created = Criterion.objects.get_or_create(
                category=category,
                parent=parent,
                excel_col=col,
                defaults={**defaults, "name": name},
            )

            if created:
                created_children += 1
            else:
                # обновим имя/тип/порядок если изменилось
                changed = False
                if obj.name != name:
                    obj.name = name
                    changed = True
                vt = defaults["value_type"]
                if obj.value_type != vt:
                    obj.value_type = vt
                    changed = True
                if obj.order != order:
                    obj.order = order
                    changed = True
                if not obj.is_active:
                    obj.is_active = True
                    changed = True
                if obj.code != f"C{col}":
                    obj.code = f"C{col}"
                    changed = True

                if changed:
                    obj.save()
                    updated_children += 1

        self.stdout.write(self.style.SUCCESS(
            f"Готово ✅ Родителей создано: {created_parents} | Полей создано: {created_children} | Обновлено: {updated_children}"
        ))
