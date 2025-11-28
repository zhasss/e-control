from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from geo.models import Region, District
from monitoring.models import School


class Command(BaseCommand):
    help = "Импорт школ из файла ОСО (лист 'Сеть') в модель School"

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Путь к файлу Excel (например, 'ОСО 07 11 25.xlsx')",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]

        # Формы собственности, которые считаем ЧАСТНЫМИ
        private_forms = {
            "Собственность граждан",
            "Собственность предприятий без государственного и иностранного участия",
            "Собственность иностранных юридических лиц",
            "Собственность международных организаций",
        }

        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Не удалось открыть файл {excel_path}: {e}")

        # Предполагаем, что нас интересует лист 'Сеть'
        if "Сеть" not in wb.sheetnames:
            raise CommandError("В файле нет листа 'Сеть'")

        ws = wb["Сеть"]

        # Читаем заголовки из первой строки
        header_row = 1
        headers = {}
        for cell in ws[header_row]:
            if cell.value is None:
                continue
            header_name = str(cell.value).strip()
            headers[header_name] = cell.column  # номер колонки (int)

        # Названия столбцов — подстроись под свои, если отличаются
        required_cols = [
            "Область",
            "Район",
            "Населенный пункт",
            "Наименование организации",
            "БИН",
            "Форма собственности",
        ]
        for col_name in required_cols:
            if col_name not in headers:
                raise CommandError(f"В листе 'Сеть' не найден столбец '{col_name}'")

        col_region = headers["Область"]
        col_district = headers["Район"]
        col_locality = headers["Населенный пункт"]
        col_name = headers["Наименование организации"]
        col_bin = headers["БИН"]
        col_ownership = headers["Форма собственности"]

        created = 0
        updated = 0
        skipped_empty_name = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=col_name).value
            if not name_val:
                skipped_empty_name += 1
                continue

            region_val = ws.cell(row=row_idx, column=col_region).value
            district_val = ws.cell(row=row_idx, column=col_district).value
            locality_val = ws.cell(row=row_idx, column=col_locality).value
            bin_val = ws.cell(row=row_idx, column=col_bin).value
            ownership_val = ws.cell(row=row_idx, column=col_ownership).value

            name_str = str(name_val).strip()
            region_str = str(region_val).strip() if region_val else ""
            district_str = str(district_val).strip() if district_val else ""
            locality_str = str(locality_val).strip() if locality_val else ""
            bin_str = str(bin_val).strip() if bin_val else ""
            ownership_str = str(ownership_val).strip() if ownership_val else ""

            if not region_str or not district_str:
                self.stdout.write(
                    self.style.WARNING(
                        f"Строка {row_idx}: нет региона/района, пропускаю"
                    )
                )
                continue

            # создаём/находим регион и район
            region, _ = Region.objects.get_or_create(name=region_str)
            district, _ = District.objects.get_or_create(
                region=region,
                name=district_str,
            )

            # определяем, частная ли школа
            is_private = ownership_str in private_forms

            # ищем школу по БИН, если есть
            school = None
            if bin_str:
                school = School.objects.filter(bin=bin_str).first()

            if school is None:
                # если БИН пустой или не нашли по нему — ищем по name+region+district
                school = School.objects.filter(
                    name=name_str,
                    region=region,
                    district=district,
                ).first()

            if school is None:
                school = School.objects.create(
                    name=name_str,
                    bin=bin_str,
                    region=region,
                    district=district,
                    locality=locality_str,
                    ownership_form=ownership_str,
                    is_private=is_private,
                )
                created += 1
            else:
                school.name = name_str
                school.locality = locality_str
                if bin_str:
                    school.bin = bin_str
                school.region = region
                school.district = district
                school.ownership_form = ownership_str
                school.is_private = is_private
                school.save()
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"Импорт завершён. "
            f"Создано: {created}, обновлено: {updated}, "
            f"пропущено без названия: {skipped_empty_name}."
        ))
