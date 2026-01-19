from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.conf import settings

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from rest_framework import viewsets, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError, PermissionDenied
from django.utils import timezone
from django.utils.dateparse import parse_date
from datetime import datetime, timedelta, time
 
from geo.models import Region, District
from .models import (
    School,
    MonitoringUser,
    InspectionType,
    CriterionCategory,
    Criterion,
    Inspection,
    InspectionResult,
)
from .serializers import (
    RegionSerializer,
    DistrictSerializer,
    SchoolSerializer,
    InspectionTypeSerializer,
    CriterionCategorySerializer,
    CriterionSerializer,
    InspectionSerializer,
    InspectionResultSerializer,
    MonitoringUserSerializer,
)
from .permissions import IsMonitoringUser


# ---------- Регионы / районы / школы ----------

class RegionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Region.objects.all().order_by("name")
    serializer_class = RegionSerializer
    permission_classes = [IsMonitoringUser]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        if user.is_superuser or not profile:
            return qs

        if profile.role == MonitoringUser.ROLE_INSPECTOR and profile.region_id:
            qs = qs.filter(id=profile.region_id)

        return qs


class DistrictViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = District.objects.select_related("region").all().order_by("name")
    serializer_class = DistrictSerializer
    permission_classes = [IsMonitoringUser]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        if user.is_superuser or not profile:
            pass
        elif profile.role == MonitoringUser.ROLE_INSPECTOR and profile.region_id:
            qs = qs.filter(region_id=profile.region_id)

        region_id = self.request.query_params.get("region")
        if region_id:
            qs = qs.filter(region_id=region_id)

        return qs


class SchoolViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = School.objects.select_related("region", "district").all()
    serializer_class = SchoolSerializer
    permission_classes = [IsMonitoringUser]
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "name_kz", "name_ru_full", "locality"]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        if user.is_superuser or not profile:
            pass
        elif profile.role == MonitoringUser.ROLE_INSPECTOR and profile.region_id:
            qs = qs.filter(region_id=profile.region_id)

        region_id = self.request.query_params.get("region")
        district_id = self.request.query_params.get("district")
        is_private = self.request.query_params.get("is_private")
        search = self.request.query_params.get("search")

        if region_id:
            qs = qs.filter(region_id=region_id)
        if district_id:
            qs = qs.filter(district_id=district_id)
        if is_private in ("true", "1"):
            qs = qs.filter(is_private=True)
        if search:
            pass

        return qs


# ---------- Критерии ----------

class InspectionTypeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = InspectionType.objects.filter(is_active=True)
    serializer_class = InspectionTypeSerializer
    permission_classes = [IsMonitoringUser]

    @action(detail=False, methods=["post"], url_path="seed-private-template")
    def seed_private_template(self, request):
        """
        1 раз запускаешь (только superuser): создаст InspectionType + Category + Criteria под шаблон +по ЧШ.xlsx
        """
        if not request.user.is_superuser:
            raise PermissionDenied("Только superuser может сидить шаблон критериев.")

        it, _ = InspectionType.objects.get_or_create(
            code="private_schools",
            defaults={"name": "Проверка частных школ", "is_active": True},
        )

        cat, _ = CriterionCategory.objects.get_or_create(
            inspection_type=it,
            name="+по ЧШ (шаблон Excel)",
            defaults={"order": 1},
        )

        # helper
        def upsert(code, name, value_type, order, excel_col=None, parent=None):
            obj, created = Criterion.objects.update_or_create(
                category=cat,
                code=code,
                defaults={
                    "name": name,
                    "value_type": value_type,
                    "order": order,
                    "excel_col": excel_col,
                    "parent": parent,
                    "is_active": True,
                },
            )
            return obj

        # --- ROOT groups (row 10 merged headers) ---
        grp_students = upsert("grp_students", "Количество учащихся", Criterion.ValueType.TEXT, 600, None, None)
        grp_director = upsert("grp_director", "Сведения по директорам", Criterion.ValueType.TEXT, 900, None, None)
        grp_teachers_new = upsert(
            "grp_teachers_new", "Категория нового формата", Criterion.ValueType.TEXT, 1600, None, None
        )
        grp_own_building = upsert("grp_own_building", "Собственное здание", Criterion.ValueType.TEXT, 2400, None, None)
        grp_rent_building = upsert("grp_rent_building", "Арендованное здание", Criterion.ValueType.TEXT, 2600, None, None)
        grp_sport = upsert("grp_sport", "Наличие спортивного зала", Criterion.ValueType.TEXT, 2800, None, None)
        grp_library = upsert("grp_library", "Наличие библиотечного фонда", Criterion.ValueType.TEXT, 3100, None, None)
        grp_cabinets = upsert("grp_cabinets", "Модифицированные кабинеты", Criterion.ValueType.TEXT, 3500, None, None)
        grp_state = upsert("grp_state", "Состояние школы", Criterion.ValueType.TEXT, 3800, None, None)

        # --- Leaf criteria (columns 2..39) ---
        upsert("has_state_order", "Размещен госзаказ", Criterion.ValueType.BOOLEAN, 200, 2)
        upsert("license_date", "Дата получения лицензии", Criterion.ValueType.DATE, 300, 3)
        upsert("project_capacity", "Проектная мощность", Criterion.ValueType.INTEGER, 400, 4)
        upsert("students_total", "Всего уч-ся", Criterion.ValueType.INTEGER, 500, 5)

        upsert("students_primary", "Уровня начального образования", Criterion.ValueType.INTEGER, 610, 6, grp_students)
        upsert("students_basic", "Уровня основного среднего образования", Criterion.ValueType.INTEGER, 620, 7, grp_students)
        upsert("students_general", "Уровня общего среднего образования", Criterion.ValueType.INTEGER, 630, 8, grp_students)

        upsert("director_education", "Образование", Criterion.ValueType.TEXT, 910, 9, grp_director)
        upsert("director_ped_category", "Пед. категория", Criterion.ValueType.TEXT, 920, 10, grp_director)
        upsert("director_manager_category", "Категория как руководителя", Criterion.ValueType.TEXT, 930, 11, grp_director)
        upsert("director_experience_total", "Общий стаж", Criterion.ValueType.INTEGER, 940, 12, grp_director)
        upsert("director_experience_position", "Стаж на данной позиции", Criterion.ValueType.INTEGER, 950, 13, grp_director)
        upsert("director_is_founder", "Директор является учредителем", Criterion.ValueType.BOOLEAN, 960, 14, grp_director)

        upsert("teachers_total", "Всего педагогов", Criterion.ValueType.INTEGER, 1500, 15)

        upsert("teachers_new_pedagog", "Педагог", Criterion.ValueType.INTEGER, 1610, 16, grp_teachers_new)
        upsert("teachers_new_moderator", "Педагог-модератор", Criterion.ValueType.INTEGER, 1620, 17, grp_teachers_new)
        upsert("teachers_new_expert", "Педагог-эксперт", Criterion.ValueType.INTEGER, 1630, 18, grp_teachers_new)
        upsert("teachers_new_researcher", "Педагог-исследователь", Criterion.ValueType.INTEGER, 1640, 19, grp_teachers_new)
        upsert("teachers_new_master", "Педагог-мастер", Criterion.ValueType.INTEGER, 1650, 20, grp_teachers_new)

        upsert("teachers_old_category", "Кол-во педагогов со старой категорией", Criterion.ValueType.INTEGER, 2100, 21)
        upsert("teachers_without_category", "Кол-во педагогов без категории", Criterion.ValueType.INTEGER, 2200, 22)
        upsert("teachers_part_time", "Кол-во совместителей", Criterion.ValueType.INTEGER, 2300, 23)

        upsert("own_building_typical", "Типовое", Criterion.ValueType.BOOLEAN, 2410, 24, grp_own_building)
        upsert("own_building_adapted", "Приспособленное", Criterion.ValueType.BOOLEAN, 2420, 25, grp_own_building)

        upsert("rented_building_typical", "Типовое", Criterion.ValueType.BOOLEAN, 2610, 26, grp_rent_building)
        upsert("rented_building_adapted", "Приспособленное", Criterion.ValueType.BOOLEAN, 2620, 27, grp_rent_building)

        upsert("sports_hall_own", "Собственный", Criterion.ValueType.BOOLEAN, 2810, 28, grp_sport)
        upsert("sports_hall_rented_small", "Арендованный (не более 1000м)", Criterion.ValueType.BOOLEAN, 2820, 29, grp_sport)
        upsert("sports_hall_rented_big", "Арендованный (более 1000м)", Criterion.ValueType.BOOLEAN, 2830, 30, grp_sport)

        upsert("library_textbooks", "Комплект учебников", Criterion.ValueType.BOOLEAN, 3110, 31, grp_library)
        upsert("library_fiction", "Художественная литература", Criterion.ValueType.BOOLEAN, 3120, 32, grp_library)

        upsert("has_computer_labs", "Наличие компьютерных классов", Criterion.ValueType.BOOLEAN, 3300, 33)
        upsert("has_video_surveillance", "Наличие видеонаблюдения", Criterion.ValueType.BOOLEAN, 3400, 34)

        upsert("cabinet_physics", "Кабинет физики", Criterion.ValueType.BOOLEAN, 3510, 35, grp_cabinets)
        upsert("cabinet_chemistry", "Кабинет химии", Criterion.ValueType.BOOLEAN, 3520, 36, grp_cabinets)
        upsert("cabinet_biology", "Кабинет биологии", Criterion.ValueType.BOOLEAN, 3530, 37, grp_cabinets)

        upsert("state_emergency", "Аварийный", Criterion.ValueType.BOOLEAN, 3810, 38, grp_state)
        upsert("state_needs_repair", "Требует ремонта", Criterion.ValueType.BOOLEAN, 3820, 39, grp_state)

        return Response({"ok": True, "inspection_type_id": it.id, "category_id": cat.id})


class CriterionCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CriterionCategory.objects.select_related("inspection_type").all()
    serializer_class = CriterionCategorySerializer
    permission_classes = [IsMonitoringUser]

    def get_queryset(self):
        qs = super().get_queryset()
        inspection_type_id = self.request.query_params.get("inspection_type")
        if inspection_type_id:
            qs = qs.filter(inspection_type_id=inspection_type_id)
        return qs


class CriterionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Criterion.objects.select_related("category", "category__inspection_type", "parent").all()
    serializer_class = CriterionSerializer
    permission_classes = [IsMonitoringUser]

    def get_queryset(self):
        qs = super().get_queryset()
        category_id = self.request.query_params.get("category")
        inspection_type_id = self.request.query_params.get("inspection_type")
        if category_id:
            qs = qs.filter(category_id=category_id)
        if inspection_type_id:
            qs = qs.filter(category__inspection_type_id=inspection_type_id)
        return qs


# ---------- Акты проверок ----------

class InspectionViewSet(viewsets.ModelViewSet):
    queryset = Inspection.objects.select_related(
        "inspection_type", "school", "region", "district", "inspector"
    ).all()
    serializer_class = InspectionSerializer
    permission_classes = [IsMonitoringUser]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        # ---- Role restrictions ----
        if user.is_superuser:
            pass
        elif profile:
            if profile.role == MonitoringUser.ROLE_INSPECTOR:
                if profile.region_id:
                    qs = qs.filter(region_id=profile.region_id)
                qs = qs.filter(inspector=user)
            elif profile.role == MonitoringUser.ROLE_CENTER:
                # Центр видит только отправленные и утвержденные
                qs = qs.filter(status__in=["completed", "approved"])
        else:
            return qs.none()

        # ---- Filters from query params ----
        qp = self.request.query_params

        school_id = qp.get("school")
        region_id = qp.get("region")
        district_id = qp.get("district")
        status_param = qp.get("status")

        inspection_type_id = qp.get("inspection_type")  # <-- ВИД ПРОВЕРКИ (id)
        year = qp.get("year")                           # <-- ГОД
        date_from = qp.get("date_from")                 # <-- ДАТА ОТПРАВКИ В ЦЕНТР (от)
        date_to = qp.get("date_to")                     # <-- ДАТА ОТПРАВКИ В ЦЕНТР (до)

        if school_id:
            qs = qs.filter(school_id=school_id)
        if region_id:
            qs = qs.filter(region_id=region_id)
        if district_id:
            qs = qs.filter(district_id=district_id)
        if status_param:
            qs = qs.filter(status=status_param)

        if inspection_type_id:
            qs = qs.filter(inspection_type_id=inspection_type_id)

        if year:
            qs = qs.filter(year=year)

        # date_from/date_to = фильтр по submitted_at (дата отправки в центр)
        # если submitted_at = null (черновик), он просто не попадет в выборку
        d_from = parse_date(date_from) if date_from else None
        d_to = parse_date(date_to) if date_to else None

        if d_from:
            dt_from = timezone.make_aware(datetime.combine(d_from, time.min))
            qs = qs.filter(submitted_at__gte=dt_from)

        if d_to:
            # inclusive: до конца дня
            dt_to = timezone.make_aware(datetime.combine(d_to + timedelta(days=1), time.min))
            qs = qs.filter(submitted_at__lt=dt_to)

        return qs

    def _can_edit_inspection(self, inspection: Inspection, user, profile) -> bool:
        if user.is_superuser:
            return True
        if not profile:
            return False
        if profile.role == MonitoringUser.ROLE_INSPECTOR:
            return inspection.inspector_id == user.id and inspection.status in ("draft", "in_progress")
        return False

    def perform_create(self, serializer):
        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        if not profile or profile.role != MonitoringUser.ROLE_INSPECTOR:
            raise PermissionDenied("Создавать акты могут только инспекторы ДОКСО.")

        inspection: Inspection = serializer.save(inspector=user)

        # регион/район подтянем из школы
        if inspection.school_id:
            if not inspection.region_id:
                inspection.region = inspection.school.region
            if not inspection.district_id:
                inspection.district = inspection.school.district
            inspection.save(update_fields=["region", "district", "updated_at"])

        # ✅ создаём результаты ТОЛЬКО по листовым критериям
        leaf_criteria = Criterion.objects.filter(
            category__inspection_type=inspection.inspection_type,
            is_active=True,
            children__isnull=True,
        ).order_by("excel_col", "order", "id")

        InspectionResult.objects.bulk_create(
            [InspectionResult(inspection=inspection, criterion=c) for c in leaf_criteria],
            ignore_conflicts=True,
        )

    def perform_update(self, serializer):
        inspection = self.get_object()
        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        if not self._can_edit_inspection(inspection, user, profile):
            raise PermissionDenied("Вы не можете редактировать этот акт проверки.")
        serializer.save()

    @action(detail=True, methods=["post"], url_path="submit")
    def submit(self, request, pk=None):
        inspection = self.get_object()
        user = request.user
        profile = getattr(user, "monitoring_profile", None)

        if not profile or profile.role != MonitoringUser.ROLE_INSPECTOR:
            raise PermissionDenied("Только инспектор ДОКСО может отправлять акт в центр.")
        if inspection.inspector_id != user.id:
            raise PermissionDenied("Вы не являетесь исполнителем по этому акту.")
        if inspection.status not in ("draft", "in_progress"):
            raise ValidationError("Акт уже отправлен в центр или утверждён.")

        total = (
            InspectionResult.objects.filter(inspection=inspection)
            .exclude(score__isnull=True)
            .aggregate(total=Sum("score"))
            .get("total")
            or 0
        )

        inspection.total_score = total
        inspection.status = "completed"
        inspection.submitted_at = timezone.now()   # ✅ ВАЖНО
        inspection.save(update_fields=["total_score", "status", "submitted_at", "updated_at"])

        serializer = self.get_serializer(inspection)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        inspection = self.get_object()
        user = request.user
        profile = getattr(user, "monitoring_profile", None)

        if not user.is_superuser:
            if not profile or profile.role != MonitoringUser.ROLE_CENTER:
                raise PermissionDenied("Только центр может утверждать акт.")
        if inspection.status != "completed":
            raise ValidationError("Утвердить можно только акт со статусом 'Отправлен в центр'.")

        inspection.status = "approved"
        inspection.approved_at = timezone.now()    # ✅ ВАЖНО
        inspection.save(update_fields=["status", "approved_at", "updated_at"])

        serializer = self.get_serializer(inspection)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        qs = self.get_queryset()

        total = qs.count()
        total_completed = qs.filter(status="completed").count()
        total_approved = qs.filter(status="approved").count()

        by_region = (
            qs.values("region_id", "region__name")
            .annotate(
                total=Count("id"),
                completed=Count("id", filter=Q(status="completed")),
                approved=Count("id", filter=Q(status="approved")),
            )
            .order_by("region__name")
        )

        data = {
            "total": total,
            "total_completed": total_completed,
            "total_approved": total_approved,
            "by_region": [
                {
                    "region_id": row["region_id"],
                    "region_name": row["region__name"],
                    "total": row["total"],
                    "completed": row["completed"],
                    "approved": row["approved"],
                }
                for row in by_region
            ],
        }
        return Response(data)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """
        Экспорт в Excel по шаблону.
        Фильтры берутся из get_queryset() (регион/район/вид/год/статус/date_from-date_to).
        """
        user = request.user
        profile = getattr(user, "monitoring_profile", None)
        if not user.is_superuser and not profile:
            raise PermissionDenied("Нет доступа к экспортy.")

        qs = (
            self.get_queryset()
            .select_related("school", "region", "district", "inspection_type")
            .prefetch_related("results__criterion")
            .order_by("region__name", "district__name", "school__name")
        )

        def yn01_to_int(v):
            if v is None:
                return None
            if isinstance(v, bool):
                return 1 if v else 0
            if isinstance(v, (int, float)):
                return 1 if int(v) == 1 else 0
            s = str(v).strip().lower()
            if s in ("1", "да", "true", "yes", "y"):
                return 1
            if s in ("0", "нет", "false", "no", "n"):
                return 0
            return None

        def result_to_excel_value(res: InspectionResult):
            if not res:
                return None
            val = res.get_value() if hasattr(res, "get_value") else None
            if res.criterion.value_type == "bool":
                return yn01_to_int(val)
            return val

        template_path = Path(settings.BASE_DIR) / "monitoring" / "templates" / "private_schools_template.xlsx"
        try:
            wb = load_workbook(template_path)
        except FileNotFoundError:
            raise ValidationError(f"Файл шаблона не найден: {template_path}")

        ws = wb.active

        if ws.cell(row=32, column=1).value == "Всего":
            ws.delete_rows(32, 1)

        START_ROW = 12
        row = START_ROW

        for insp in qs:
            school = insp.school
            if not school:
                continue

            ws.cell(row=row, column=1, value=getattr(school, "name", ""))

            for r in insp.results.all():
                col = getattr(r.criterion, "excel_col", None)
                if not col or col == 1:
                    continue
                ws.cell(row=row, column=col, value=result_to_excel_value(r))

            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="private_schools.xlsx"'
        return response
# ---------- Результаты по критериям ----------

class InspectionResultViewSet(viewsets.ModelViewSet):
    queryset = InspectionResult.objects.select_related("inspection", "criterion", "criterion__category").all()
    serializer_class = InspectionResultSerializer
    permission_classes = [IsMonitoringUser]

    def get_queryset(self):
        qs = super().get_queryset()
        inspection_id = self.request.query_params.get("inspection")
        if inspection_id:
            qs = qs.filter(inspection_id=inspection_id)

        user = self.request.user
        profile = getattr(user, "monitoring_profile", None)

        if user.is_superuser or not profile:
            return qs

        if profile.role == MonitoringUser.ROLE_INSPECTOR and profile.region_id:
            qs = qs.filter(inspection__region_id=profile.region_id)
        elif profile.role == MonitoringUser.ROLE_CENTER:
            qs = qs.filter(inspection__status__in=["completed", "approved"])

        return qs

    def _ensure_editable(self, inspection: Inspection, user):
        profile = getattr(user, "monitoring_profile", None)

        if user.is_superuser:
            return
        if not profile:
            raise PermissionDenied("Нет доступа к мониторингу.")
        if inspection.status not in ("draft", "in_progress"):
            raise ValidationError("Проверка уже отправлена в центр и не может быть изменена.")
        if profile.role != MonitoringUser.ROLE_INSPECTOR or inspection.inspector_id != user.id:
            raise PermissionDenied("Вы не можете редактировать результаты по этому акту.")

    def perform_create(self, serializer):
        inspection_id = self.request.data.get("inspection")
        inspection = get_object_or_404(Inspection, pk=inspection_id)
        self._ensure_editable(inspection, self.request.user)
        serializer.save(inspection=inspection)

    def update(self, request, *args, **kwargs):
        instance: InspectionResult = self.get_object()
        self._ensure_editable(instance.inspection, request.user)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        instance: InspectionResult = self.get_object()
        self._ensure_editable(instance.inspection, request.user)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance: InspectionResult = self.get_object()
        self._ensure_editable(instance.inspection, request.user)
        return super().destroy(request, *args, **kwargs)


# ---------- /api/monitoring/me ----------

class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        profile = getattr(user, "monitoring_profile", None)

        data = {
            "user": {
                "id": user.id,
                "username": user.username,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "is_superuser": user.is_superuser,
            },
            "monitoring_profile": None,
        }

        if profile:
            data["monitoring_profile"] = MonitoringUserSerializer(profile).data

        return Response(data)
